from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border , Side, PatternFill, Protection
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.protection import SheetProtection
from openpyxl.chart import BarChart, Reference

import string
import copy
import pickle

import requests
from pymongo import MongoClient

import matplotlib.pyplot as plt

import Levenshtein
import jaro
import pandas as pd
import numpy as np
import json
import typing
import hashlib
from bson.objectid import ObjectId


# Create a hashlib object for the SHA-256 hash
hash_object = hashlib.sha256()

ONEDRIVE_STORAGE_HASH = "b!AexAOmJBX0GFrFCDPly37vK7ahDKKlBHgbMmpv0CDCXt3nRYonxISZkyQ12XYZPz"
BANQUE_STORAGE_HASH = "b!sYa62uHHtUqTlMT05hXNNDimI6dJj1xDlCQOeoCbtQ2y7z_UU_D9QKTSaBBIB995"
GOLD_DOCUMENT_FOLDER = [ONEDRIVE_STORAGE_HASH, "Banque documents", "Filed"]
PREDICTION_DOCUMENT_FOLDER = [BANQUE_STORAGE_HASH, "Clients"]

class MyMongoClient :
    def __init__(self,
                 db,
        MONGO_HOST = 'localhost',
        MONGO_PORT = 56788,
        MONGO_USER = 'root',
        MONGO_PASSWORD = "7pk_ht*JvMjUpXFF!qxE" # TODO: Remove this password from the code
        ):
        self.db = db
        self.mongo_client = MongoClient(f'mongodb://{MONGO_USER}:{MONGO_PASSWORD}@{MONGO_HOST}:{MONGO_PORT}')
        self.all_docs = None

    def get_all_field_of_documents(self):
        cursor = self.mongo_client[self.db]['documents']
        documents = cursor.find({})
        fields = set()
        for i, doc in enumerate(documents):
            for field in doc['information']:
                fields.add(field)
        return fields
    
    def get_str_to_match(self, fields):
        """
        return a dictionary with 
        value : the value
        field : the field name 
        doc_hash : the hash of the docuement it is in 
        """
        cursor = self.mongo_client[self.db]['documents']
        documents = cursor.find({})

        str_to_match = []
        for i, doc in enumerate(documents):
            for field in fields:
                if field in doc['information'] :#'sign' in str(doc) or 'fournisseur' in str(doc) or 'destinataire' in str(doc):
                    str_to_match.append(
                        {"value": doc['information'][field], 
                        'field' : field, 
                        'doc_hash': doc['_id']} 
                    )
        return str_to_match
    
    def get_values_to_match(self, fields):
        # We load all the stings that were not entities
        all_values = self.get_str_to_match(self.db, fields)

        df = pd.DataFrame(all_values)
        df['doc_field'] = df.apply(lambda row : (row['field'], row['doc_hash']), axis = 1)
        df_grouped = pd.DataFrame(df.groupby('value').doc_field.agg(set)).reset_index()
        out = []
        for idx, row in df_grouped.iterrows():
            out.append({
                'value' : row['value'],
                'docs_hashs' : row['doc_field']
                })
        return out

    def get_all_entities(self):
        cursor = self.mongo_client[self.db]['entities']
        entities = cursor.find({})
        return entities
    
    def get_all_documents(self) :
        if not self.all_docs :
            cursor = self.mongo_client[self.db]['documents']
            self.all_docs = list(cursor.find({}))
        return self.all_docs
    
    def get_docs_labels(self, path_storage):
        all_docs = self.get_all_documents()
        filtered_data = []
        file_names = []
        for item in all_docs:
            for file in item.get("files", [{}]) :
                if file.get("fullPath")[:len(path_storage)] == path_storage:
                    filtered_data.append(item)
                    file_names.append(file['fileName'])
        # filtered_data = [item for item in all_docs 
        #                  if item.get("files", [{}])[0].get("fullPath")[:len(path_storage)] == path_storage]
        # print([doc['files'][0]['fileName'] for doc in filtered_data[0:3]])
        labels = {doc['_id'] : {
            label['name'] : label['value'] for label in doc['labels']}
                     for doc in filtered_data}
        return labels, file_names
    
    def get_results(self):
        gold_docs_labels, gold_names = load_gold_labels()
        pred_docs_labels, pred_names = self.get_docs_labels(PREDICTION_DOCUMENT_FOLDER)
        missing, wrong, right, total = 0,0,0,0
        acc_by_doc = {}
        # print(gold_names)
        # print(pred_names)
        # print(len(gold_docs_labels),len(pred_docs_labels))
        # print(set(gold_docs_labels.keys()) & set(pred_docs_labels.keys()))

        for i, doc_labels in enumerate(gold_docs_labels) :
            if doc_labels not in pred_docs_labels:
                print(f"{gold_names[i]} is not in prediction label")
                continue
            doc_right, doc_total = 0,len(gold_docs_labels[doc_labels])
            for key in gold_docs_labels[doc_labels] :
                if key in ['language', 'filename date']:
                    doc_total -= 1
                elif key in pred_docs_labels[doc_labels]:
                    if pred_docs_labels[doc_labels][key] == gold_docs_labels[doc_labels][key] :
                        doc_right +=1
                    else :
                        wrong +=1 
                        print(f"{gold_names[i]} Wrong field {key}. Found {pred_docs_labels[doc_labels][key]} instead of {gold_docs_labels[doc_labels][key]}")
                    
                else :
                    print(f"{gold_names[i]}Missing field : {key}")
                    missing += 1
            right += doc_right
            total += doc_total
            acc_by_doc[doc_labels] = doc_right / doc_total
        if not total :
            return "No document in common"
        general_accuracy = right / total
        others = {
            "missing" : missing,
            "wrong" : wrong,
            "right" : right,
            "total" : total
        }
        return general_accuracy, acc_by_doc, others 
        
    def get_hash_of_doc(self, doc_id):
        hash_object = hashlib.sha256()
        input_string = list(self.mongo_client[self.db]['contents'].find({'_id' : doc_id}))[0]['content']
        # Update the hash object with the bytes of your string
        hash_object.update(input_string.encode('utf-8'))

        # Get the hexadecimal representation of the hash
        return hash_object.hexdigest()

def load_gold_labels():
    with open("./saves/2023-11-01-gold_labels_filenames_clean", 'rb') as f :
        res = pickle.load(f)
    return res['labels'], res['file_names']


class EntityMatching :
    def __init__(self,mongo_client : MyMongoClient, nb_proposition :int = 4, confidence_threshold : int = 0.6, can_create_entity : bool = True) :
        self.NB_PROPOSITIONS = nb_proposition
        self.CONFIDENCE_THRESH = confidence_threshold
        self.mongo_client = mongo_client
        self.can_create_entity = can_create_entity
        self.excel_shift = 2 if self.can_create_entity else 1   # add 1 for create_new_entity add 1 for ignore
    
    def propose_best_matches(self,  db, fields, excel_file_path):
        # We load all the values that need to be match alongside the documents they are in and in which field
        val_to_match = self.mongo_client.get_values_to_match(db, fields)

        # We take all the entities of the organisation
        entities = list(self.mongo_client.get_all_entities(db))

        #Get all the best matches for all the strings that are not entities
        all_best_matches = self.get_all_best_matches(val_to_match, entities)
        self.all_best_matches = all_best_matches

        #Write an excel for the client
        self.write_excel(all_best_matches, excel_file_path)

        #Add styling to the excel
        self.decorate_excel(excel_file_path, color = "F2F2F2")
        
        self.plot_confidence_in_excel(excel_file_path)


    def get_lev_jaro_scores(self, entity_names, testing_name):    
        lev_scores = [Levenshtein.ratio(testing_name, ent) for ent in entity_names]
        jaro_scores = [jaro.jaro_metric(testing_name, ent) for ent in entity_names]
        jaro_wink_scores = [jaro.jaro_winkler_metric(testing_name, ent) for ent in entity_names]

        return pd.DataFrame({'names' : entity_names,
                    'lev_score' : lev_scores,
                    'jaro_scores' : jaro_scores,
                    'jaro_wink_scores' : jaro_wink_scores})

    def get_jaro_scores(self, entities, testing_name):
        entity_names = [ent['information']['Display name'] for ent in entities]
        entity_id    = [ent['_id'] for ent in entities]
        jaro_scores  = [jaro.jaro_winkler_metric(testing_name, ent) for ent in entity_names]
        return pd.DataFrame({
            '_id' : entity_id,
            'names' : entity_names,
            'jaro_scores' : jaro_scores
        })

    def get_best_matches(self, entities, testing_name):
        best_matches = self.get_jaro_scores(entities, testing_name).sort_values('jaro_scores', ascending = False)[:self.NB_PROPOSITIONS]
        out = []
        for idx, row in best_matches.iterrows() : 
            out.append({
                '_id' : str(row['_id']),
                'name' : row['names'],
                'jaro_score' : row['jaro_scores']
            })
        return out
    


    def get_all_best_matches(self, values_to_match, entities) :
        scores = []
        for value_dict in values_to_match :
            value_dict["propositions"] = self.get_best_matches(entities, value_dict["value"])
            scores.extend([prop['jaro_score'] for prop in value_dict["propositions"]])
            value_dict['confidence'] = 1 if value_dict['propositions'][0]['jaro_score']  == 1 else  (2*value_dict['propositions'][0]['jaro_score'] - value_dict['propositions'][1]['jaro_score'])
        
        confidences = [val['confidence'] for val in values_to_match]
        min_, max_ = min(scores), max(confidences)
        for val in values_to_match :
            val['confidence'] =  1 if val['confidence']== 1 else (val['confidence'] - min_)/(max_+ 0.05 - min_)
        values_to_match = sorted(values_to_match, key = lambda value_dict : -value_dict['confidence'])
        return values_to_match
    
    
    def plot_confidence(self):
        # Sample data
        data = [val['confidence'] for val in self.all_best_matches]
        print(data)
        # Custom bucket edges
        buckets = [0 ,0.15, 0.25, 0.35, 0.45, 0.55, 0.65, 0.75, 0.85, 0.95, 0.99, 1.01]
        midpoints = np.array(buckets[:-1]) + np.diff(buckets) / 2
        # Create the histogram
        plt.hist(data, bins=buckets, width=0.05, edgecolor='black', align = 'mid')

        plt.xticks(midpoints)

        # Customize the histogram
        plt.xlabel('Confidence')
        plt.ylabel('Frequency')
        plt.title('Distribution of the confidence')
        plt.show()

    def plot_confidence_in_excel(self, excel_file_path):
        # Load the Excel file
        workbook = load_workbook(excel_file_path)

        # Select the desired sheet
        sheet_matches = workbook['Matches']
        sheet = workbook['Confidence']  # Replace 'Sheet1' with the actual sheet name
        values = [float(cell.value[13:]) for i,cell in enumerate(sheet_matches[string.ascii_uppercase[self.NB_PROPOSITIONS]]) if i % 4 + int(self.can_create_entity)== 1]
        # #Hise the columns with the ids
        # for col in string.ascii_uppercase[:2]:
        #     sheet.column_dimensions[col].hidden = True


        max_val = 1.01
        bin_size = 0.1
        # bins = np.arange(-0.1, max_val+0.01, bin_size)
        # labels = np.arange(0 , max_val , bin_size)
        bins = np.arange(-0.0001, max_val+0.11, bin_size)
        labels = np.arange(0 , max_val , bin_size)
        # Assign values to each group
        df = pd.DataFrame()
        df['confidence'] = values
        df['confidence_hist'] = pd.cut(df['confidence'], bins, labels=labels)
        df = pd.DataFrame(df.groupby('confidence_hist').confidence.count()).reset_index()
        # Add the data to the sheet

        for i, row in df.iterrows():
            sheet.cell(row=i+2, column=4, value=row['confidence_hist'])
            sheet.cell(row=i+2, column=5, value=row['confidence'])

        # Create a Histogram chart
        chart = BarChart()
        chart.title = "Distribution du degré de confidence"
        chart.x_axis.title = "Degré de confiance"
        chart.y_axis.title = "Quantité"

        # Remove the legend
        # chart.legend = None

        # Set the data range for the chart
        data_range = Reference(sheet, min_col=5, min_row=1, max_row=len(df)+1)
        categories = Reference(sheet, min_col=4, min_row=2, max_row=len(df)+1)
        chart.add_data(data_range, titles_from_data=True)
        chart.set_categories(categories)


        chart.height = 12
        chart.width = 18
        # Add the chart to the sheet
        sheet.add_chart(chart, "D1")

        # Save the workbook
        workbook.save(excel_file_path)


    def write_excel(self, strings, excel_file_path):
        out = []
        base_line = ['']* (2* self.NB_PROPOSITIONS +self.excel_shift +1)


        line = base_line.copy()
        line[0] = f"MATCHING ENTITIES"
        line[-1] = "title"
        out.append(line)
        for value_dict in strings :
            value = value_dict['value']
            propositions = value_dict['propositions']
            docs_hashs = value_dict['docs_hashs']
            confidence = value_dict['confidence']
            
            
            #Append a line for the question 
            line = base_line.copy()
            line[0] = f"Name to be linked :"
            line[1] = value
            line[2] = f"trouvé dans {set([field for field, hash in docs_hashs])}"
            line[3] = f"Confidence : {round(confidence,2)}"
            line[self.NB_PROPOSITIONS+self.excel_shift+1] = docs_hashs
            line[-1] = "question"
            out.append(line)

            #Append a line for the answer 
            line = base_line.copy()
            line[0] = f"Answer :"
            line[1] = 1 if confidence >self.CONFIDENCE_THRESH  else self.NB_PROPOSITIONS+1 
            line[-1] = "answer"
            out.append(line)

            #Append a line with the possibilities and the ids
            line = base_line.copy()
            for i, prop in enumerate(propositions) :
                line[i] = f"{i+1}. '{prop['name']}'"
                line[i+self.NB_PROPOSITIONS+self.excel_shift] = prop['_id']

            line[self.NB_PROPOSITIONS] = f"{self.NB_PROPOSITIONS+1}. Ignore value"
            if self.can_create_entity : 
                line[self.NB_PROPOSITIONS+self.excel_shift-1] = f"{self.NB_PROPOSITIONS+self.excel_shift}. Create new Entity"

            line[-1] = "possible_answers"
            out.append(line)

            #Append a line with the possibilities and the ids
            if self.can_create_entity :
                line = base_line.copy()
                line[-1] = "entity_type"
                out.append(line)

            #Append a separation line
            line = base_line.copy()
            line[-1] = "separation"
            out.append(line)
        pd.DataFrame(out).to_excel(excel_file_path, index = False, header = False)


    def decorate_excel(self, excel_file_path, new_name = "Matches", color = None) :
        basic_size = 12
        fonts = {
            'title' : Font(size = basic_size+6, bold = True),
            'question' : [
                Font(size = basic_size, italic = True),
                Font(size = basic_size, bold = True),
            ],
            'dropdown_type' : Font(size = basic_size, italic = True),
            'answer' : [
                Font(size = basic_size, italic = True),
                Font(size = basic_size, bold = True)
            ],
            'possible_answers' : Font(size = basic_size),
            'separation' : Font(size = basic_size)
        }

        border = lambda size : Border(
            left=Side(border_style=size, color='000000'),
            right=Side(border_style=size, color='000000'),
            top=Side(border_style=size, color='000000'),
            bottom=Side(border_style=size, color='000000')
        )

        # Load the Excel file
        workbook = load_workbook(excel_file_path)

        # Select the desired sheet
        sheet = workbook['Sheet1']  # Replace 'Sheet1' with the actual sheet name
        sheet.title = new_name

        #Se tthe width of the columns 
        for col in string.ascii_uppercase[:self.NB_PROPOSITIONS+1+int(self.can_create_entity)]:
            sheet.column_dimensions[col].width = 40

        #Hise the columns with the ids
        for col in string.ascii_uppercase[self.NB_PROPOSITIONS+1+int(self.can_create_entity):2*(self.NB_PROPOSITIONS+self.excel_shift)]:
            sheet.column_dimensions[col].hidden = True

        # Iterate over all cells in the sheet and enable text wrapping
        for row in sheet.iter_rows():
            for i in range(self.NB_PROPOSITIONS+self.excel_shift):
                row[i].alignment = Alignment(wrapText=True)

        # Define the values you want to allow   
        # allowed_values  = ["Type : Personne|Client", "Type : Personne|Employé", "Type : Personne|Employé Client", "Type : Entreprise|Client", "Type : Entreprise|Fournisseur"]
        allowed_values  = ["Type : Personne", "Type : Entreprise"]

        # Create a DataValidation object
        data_val_type = DataValidation(type="list", formula1=f'"{",".join(allowed_values)}"')

        # Set other properties of the DataValidation object
        data_val_type.showDropDown = False
        data_val_type.errorTitle = "Valeur invalide"
        data_val_type.error = "Si vous voulez créer une entité, choisissez une valeur."
        data_val_type.promptTitle = "Choisissez le type de l'entité."
        data_val_type.prompt = "Si vous décidez de créer une nouvelle entité, choisissez un type d'entité que vous voulez créer."
        data_val_type.font = fonts['dropdown_type']
        sheet.add_data_validation(data_val_type)



        data_val_answer = DataValidation(type="list", formula1=f'"{",".join([str(i) for i in range(1,self.NB_PROPOSITIONS+self.excel_shift+1)])}"')
        # data_val_answer = DataValidation(type="list", formula1=f'"{",".join(range(1,self.NB_PROPOSITIONS+2))}"')

        # Set other properties of the DataValidation object
        data_val_answer.showDropDown = False
        data_val_answer.errorTitle = "Valeur invalide"
        data_val_answer.error = f"Choisissez une valeur entre 1 et {self.NB_PROPOSITIONS+self.excel_shift}"
        data_val_answer.promptTitle = ""
        data_val_answer.prompt = ""
        data_val_answer.font = fonts["answer"][1]
        sheet.add_data_validation(data_val_answer)

        # Set sheet protection options
        protection = SheetProtection(sheet)
        protection.password = ""  # Optional: Set a password to unlock the sheet
        protection.sheet = True
        protection.selectLockedCells = False
        protection.selectUnlockedCells = False
        protection.formatCells = True
        protection.formatColumns = True
        protection.formatRows = True
        protection.insertColumns = True
        protection.insertRows = True
        protection.insertHyperlinks = True
        protection.deleteColumns = True
        protection.deleteRows = True
        protection.sort = True
        protection.autoFilter = True
        protection.pivotTables = True
        protection.editObjects = False

        # Enable sheet protection
        sheet.protection = protection


        #Add the font to the different rows
        for i, row in enumerate(sheet.iter_rows()):  # Start from the second row, assuming the first row is a header
        # Si c'est le nom du genre
            i = i+1
            line_desc = row[-1].value

            if line_desc == 'field_name' :
                row[0].font = fonts[line_desc]

            elif line_desc == 'title' :
                row[0].font = fonts[line_desc]
            elif line_desc == 'question' :
                row[0].font = fonts[line_desc][0]
                row[0].alignment = Alignment(horizontal='right')
                row[1].font = fonts[line_desc][1]
                row[2].font = fonts[line_desc][0]
                row[3].font = fonts[line_desc][0]

            elif line_desc == 'answer' :
                row[0].font = fonts[line_desc][0]
                row[0].alignment = Alignment(horizontal='right')
                row[1].font = fonts[line_desc][1]
                row[1].alignment = Alignment(horizontal='center')
                row[1].border = border('thick')
                row[1].protection = Protection(locked=False)
                data_val_answer.ranges.add(row[1].coordinate)

            elif line_desc == 'entity_type' and self.can_create_entity:
                row[self.NB_PROPOSITIONS+self.excel_shift-1].value = allowed_values[0]
                row[self.NB_PROPOSITIONS+self.excel_shift-1].font = Font(size = basic_size -2, italic =True)
                data_val_type.ranges.add(row[self.NB_PROPOSITIONS+self.excel_shift-1].coordinate)
                row[self.NB_PROPOSITIONS+self.excel_shift-1].border = border('thin')
                row[self.NB_PROPOSITIONS+self.excel_shift-1].protection = Protection(locked=False)

            elif line_desc == 'possible_answers' :
                for i in range(self.NB_PROPOSITIONS+self.excel_shift) :
                    row[i].font = fonts[line_desc]
                    row[i].alignment = Alignment(horizontal='center')

            elif line_desc == 'separation' :
                row[0].font = fonts[line_desc]

            if line_desc not in ["field_name", 'separation'] and color != None :
                for i in range(self.NB_PROPOSITIONS+1+int(self.can_create_entity)) :
                    row[i].fill = PatternFill(start_color=color, end_color=color,
                                            fill_type = "solid")
        
        # To adjust the columns to the field of the column
        # for column_cells in sheet.columns:
        #     max_length = 0
        #     column = column_cells[0].column_letter
        #     for cell in column_cells:
        #         if cell.value:
        #             cell_length = len(str(cell.value))
        #             if cell_length > max_length:
        #                 max_length = cell_length
        #     adjusted_width = (max_length +2 ) * 1.02  # Adding some padding and a scaling factor
        #     sheet.column_dimensions[column].width = adjusted_width

        
        # Create a new sheet
        new_sheet = workbook.create_sheet(title='Confidence')

        workbook.save(excel_file_path)


    def process_excel_filled(self, excel_file_path, processed_file_path):
        out = []
        wb = load_workbook(excel_file_path)

        # Get the sheet
        sheet = wb['Matches'] 

        # Iterate through the rows
        current_matching = {}
        for row in sheet.iter_rows():
            line_desc = row[-1].value
            if line_desc == 'separation' :
                current_matching = current_matching.copy()
            elif line_desc == 'question' :
                current_matching['value'] = row[1].value
                current_matching['doc_hashs'] = row[self.NB_PROPOSITIONS+self.excel_shift+1].value

            elif line_desc == 'answer' :
                # If the field was not answered we create a new entity
                current_matching['answer'] = row[1].value

            elif line_desc == 'entity_type' :
                # If the field was not answered we create a new entity
                current_matching['entity_type'] =  row[self.NB_PROPOSITIONS+ self.excel_shift-1].value[7:]

            elif line_desc == 'possible_answers' :
                if current_matching['answer'] > self.NB_PROPOSITIONS : #If it is ignore or create entity
                    current_matching['answer_id'] = ""
                    current_matching['answer_name'] = row[current_matching['answer']-1 ].value[3:]
                else :
                    current_matching['answer_id'] = row[current_matching['answer']+self.NB_PROPOSITIONS+ self.excel_shift-1].value
                    current_matching['answer_name'] = row[current_matching['answer']-1 ].value[4:-1]
                
                if current_matching['answer'] != self.NB_PROPOSITIONS +1: # Si on a pas la colonne ignore 
                    out.append(current_matching)
        if self.can_create_entity :
            pd.DataFrame(out)[['value', 'answer_id', 'answer_name', 'doc_hashs', 'entity_type']].to_csv(processed_file_path)
        else :
            pd.DataFrame(out)[['value', 'answer_id', 'answer_name', 'doc_hashs']].to_csv(processed_file_path)
    
random_people = [
    {"name": "Johnson", "surname": "Ethan"},
    {"name": "Smith", "surname": "Olivia"},
    {"name": "Thompson", "surname": "Liam"},
    {"name": "Wilson", "surname": "Ava"},
    {"name": "Anderson", "surname": "Noah"},
    {"name": "Martinez", "surname": "Isabella"},
    {"name": "Davis", "surname": "Sophia"},
    {"name": "Clark", "surname": "Mason"},
    {"name": "Taylor", "surname": "Mia"},
    {"name": "Walker", "surname": "Lucas"},
    {"name": "Brown", "surname": "Emma"},
    {"name": "Miller", "surname": "Jacob"},
    {"name": "Moore", "surname": "Charlotte"},
    {"name": "Wilson", "surname": "William"},
    {"name": "Harris", "surname": "Ava"},
    {"name": "Turner", "surname": "Liam"},
    {"name": "Carter", "surname": "Olivia"},
    {"name": "Baker", "surname": "Sophia"},
    {"name": "Bell", "surname": "Michael"},
    {"name": "Gonzalez", "surname": "Emily"},
    {"name": "Wright", "surname": "Alexander"},
    {"name": "Parker", "surname": "Madison"},
    {"name": "Roberts", "surname": "Daniel"},
    {"name": "Green", "surname": "Abigail"},
    {"name": "Hall", "surname": "David"},
    {"name": "Mitchell", "surname": "Grace"},
    {"name": "Young", "surname": "Andrew"},
    {"name": "Turner", "surname": "Ella"},
    {"name": "King", "surname": "Benjamin"},
    {"name": "Scott", "surname": "Sofia"},
    {"name": "Phillips", "surname": "James"},
    {"name": "Ward", "surname": "Emily"},
    {"name": "Stewart", "surname": "Jacob"},
    {"name": "Morris", "surname": "Elizabeth"},
    {"name": "Jenkins", "surname": "Matthew"},
    {"name": "Adams", "surname": "Amelia"},
    {"name": "Russell", "surname": "William"},
    {"name": "Lee", "surname": "Samantha"},
    {"name": "Carter", "surname": "Logan"},
    {"name": "Wright", "surname": "Sofia"},
    {"name": "Clark", "surname": "Henry"},
    {"name": "Foster", "surname": "Scarlett"},
    {"name": "Turner", "surname": "Ryan"},
    {"name": "Barnes", "surname": "Grace"},
    {"name": "Mitchell", "surname": "John"},
    {"name": "Flores", "surname": "Avery"},
    {"name": "Bennett", "surname": "Chloe"},
    {"name": "Perry", "surname": "Christopher"},
    {"name": "Bell", "surname": "Victoria"},
    {"name": "Martin", "surname": "Daniel"},
    {"name": "Gonzalez", "surname": "Zoe"}]

#Tests 
# entity_n_s = [f"{ent['name']} {ent['surname']}" for ent in random_people][-10:]
# get_lev_jaro_scores(entity_n_s, "Christopher Bennett").sort_values('jaro_scores')

# entity_n_s = [f"{ent['surname']} {ent['name']}" for ent in random_people][-10:]
# get_lev_jaro_scores(entity_n_s, "Bennett Christopher").sort_values('jaro_scores')